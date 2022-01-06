# -*- coding: utf-8 -*-
from scrapy import Spider, Request
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import csv, xlsxwriter, xlrd, xlwt
from xlutils.copy import copy
import openpyxl
def RepresentsInt(s):
    try:
        int(s)
        return True
    except ValueError:
        return False

class inmet_gov_brSpider(Spider):
    name = "inmet_gov_br"
    start_url = 'https://www.but.fr/mobilier/canape-fauteuil/toute-l-offre-canape-et-fauteuil/index-c11592.html'
    domain1 = 'http://www.inmet.gov.br'

    driver = None
    conn = None
    total_message_count = 0

    ids = []
    all_date_data = []
    all_stations = []

    start_date = ''
    end_date = ''
    file_path = ''

    workbook_to_write = None
    sheet_to_write = None

    wb = None


    now_row_count = 1

    def start_requests(self):
        yield Request(self.start_url)

    def parse(self, response):

        email = 'alvaromollica@gmail.com'
        password = '26d0vh1u'
        base_url = 'http://www.inmet.gov.br/projetos/rede/pesquisa/gera_serie_txt.php?&mRelEstacao={}&btnProcesso=serie&mRelDtInicio={}&mRelDtFim={}&mAtributos=,,,,,,,,,,1,,,,,'

        chrome_options = Options()
        chrome_options.add_argument("window-size=1500,1500")

        self.driver = webdriver.Chrome("chromedriver.exe", options=chrome_options)
        self.driver.set_page_load_timeout(300)

        # login #####################################
        self.driver.get('http://www.inmet.gov.br/projetos/rede/pesquisa/inicio.php')

        # login_url = self.driver.find_element_by_xpath('//div[@id="fale_conosco_main"]/div/iframe').get_attribute('src')
        # self.driver.get(login_url)
        #
        # market = self.driver.find_element_by_xpath('//form[@id="login"]/input"]')
        # actions = ActionChains(self.driver)
        # actions.move_to_element(market).perform()


        username_input = self.driver.find_element_by_xpath('//input[@type="text"]')
        username_input.send_keys(email)

        password_input = self.driver.find_element_by_xpath('//input[@type="password"]')
        password_input.send_keys(password)

        self.driver.find_element_by_xpath('//input[@value=" Acessar "]').click()

        # book_to_read = xlrd.open_workbook(self.file_path)
        # sheet_to_read = book_to_read.sheet_by_index(0)
        self.wb = openpyxl.load_workbook(self.file_path)

        w_sheet = self.wb.active

        # rb = xlrd.open_workbook(self.file_path)
        # wb = copy(rb)
        # w_sheet = wb.get_sheet(0)

        for m, id_val in enumerate(self.ids):

            self.driver.get(base_url.format(id_val, self.start_date, self.end_date))
            data = self.driver.page_source.split('Estacao;Data;Hora;Precipitacao;')[-1].split('</pre>')[0].split('\n')

            if 'Estacao;Data;Hora;Precipitacao;' not in self.driver.page_source:
                print('\n########################################')
                print("The data is not existing. Run again")
                print('########################################\n')

                self.driver.close()

                self.driver = webdriver.Chrome("chromedriver.exe", options=chrome_options)
                self.driver.set_page_load_timeout(300)

                # login #####################################
                self.driver.get('http://www.inmet.gov.br/projetos/rede/pesquisa/inicio.php')

                # login_url = self.driver.find_element_by_xpath('//div[@id="fale_conosco_main"]/div/iframe').get_attribute('src')
                # self.driver.get(login_url)
                #
                # market = self.driver.find_element_by_xpath('//form[@id="login"]/input"]')
                # actions = ActionChains(self.driver)
                # actions.move_to_element(market).perform()


                username_input = self.driver.find_element_by_xpath('//input[@type="text"]')
                username_input.send_keys(email)

                password_input = self.driver.find_element_by_xpath('//input[@type="password"]')
                password_input.send_keys(password)

                self.driver.find_element_by_xpath('//input[@value=" Acessar "]').click()
                self.driver.get(base_url.format(id_val, self.start_date, self.end_date))
                data = self.driver.page_source.split('Estacao;Data;Hora;Precipitacao;')[-1].split('</pre>')[0].split('\n')
                if 'Estacao;Data;Hora;Precipitacao;' not in self.driver.page_source:
                    print('\n########################################')
                    print("The data is not existing. Run again")
                    print('########################################\n')

                # continue

            try:
                col_of_station = self.ids.index(id_val) + 2
            except:
                col_of_station = 2

            statiiion_name = self.driver.page_source.split('Estação           :')[-1].split('\n')[0].strip()
            w_sheet.cell(row=1, column=col_of_station).value = statiiion_name

            count_to_add = 1
            for n, data_row in enumerate(data):
                data_row = data_row.strip()
                if not data_row:
                    continue
                temps = data_row.split(';')
                try:
                    date_data = temps[1]
                    rain_fall = temps[3]
                except:
                    continue


                try:
                    row_of_date = self.all_date_data.index(date_data) + 2
                except Exception as e:
                    row_of_date = self.now_row_count + count_to_add
                    # w_sheet.cell(row=row_of_date, column=1).value = str(date_data)
                    self.all_date_data.append(str(date_data))
                    # w_sheet.write(row_of_
                count_to_add += 1

                # w_sheet.write(row_of_date, col_of_station, str(rain_fall))#write station id on first row
                # w_sheet.cell(row=row_of_date, column=col_of_station).value = 'dsfsdaf'
                w_sheet.cell(row=row_of_date, column=col_of_station).value = str(rain_fall)
            print('\n########################################')
            print(str(m + 1) + '. ' + statiiion_name + ' ------ Completed')
            print('########################################\n')

            # if m == 10:
            #     break
        self.wb.save(self.file_path)